
# # 根據你提到的需求，以下是學習和教學如何從資料庫串接（MSSQL 和 MySQL）並處理來自各種格式的數據（Excel、Word、CSV、TXT）的內容大綱：

# # 1. 資料庫串接
# # MSSQL（只讀）：

# # 安裝與設定：安裝 SQL Server 驅動（如 pyodbc 或 pymssql）。
# # 連接 MSSQL：
# # 使用 Python 連接 MSSQL，示範如何進行只讀操作。
# # 使用基本的 SQL 查詢來檢索數據。
# # 連接範例：

# import pyodbc

# conn = pyodbc.connect('DRIVER={SQL Server};SERVER=your_server;DATABASE=your_database;UID=your_user;PWD=your_password')
# cursor = conn.cursor()
# cursor.execute("SELECT * FROM your_table")
# for row in cursor:
#     print(row)
# conn.close()
# # MySQL（測試與實驗）：

# # 安裝與設定：安裝 MySQL Connector/Python。
# # 建立與連接 MySQL：
# # 建立新的資料庫與資料表。
# # 使用 Python 連接 MySQL，進行資料的插入、更新、刪除操作。
# # 連接範例：

# import mysql.connector

# conn = mysql.connector.connect(
#     host="10.96.196.36",
#     user="stan",
#     password="Ss31668984",
#     database="stan_test"
# )
# cursor = conn.cursor()
# cursor.execute("INSERT INTO speedtest_results (column1, column2) VALUES (%s, %s)", (value1, value2))
# conn.commit()
# cursor.close()
# conn.close()
# # 2. 資料庫建立與管理
# # 自建資料庫：
# # 在 MySQL 上建立資料庫及表格，規劃資料庫的結構和關聯。
# # 使用 SQL 語句來創建表格，定義主鍵、外鍵、索引等。
# # 資料上傳至 MySQL：
# # 從 Excel、CSV、TXT 等文件中讀取數據，並將其插入到 MySQL 資料庫中。
# # 範例代碼：

import pandas as pd
import mysql.connector

# # 讀取 CSV 文件
# df = pd.read_csv('your_file.csv')

# # 連接到 MySQL
# conn = mysql.connector.connect(
#     host="10.96.196.36",
#     user="stan",
#     password="Ss31668984",
#     database="stan_test"
# )
# cursor = conn.cursor()

# # 插入數據
# for i, row in df.iterrows():
#     cursor.execute("INSERT INTO your_table (column1, column2) VALUES (%s, %s)", (row['col1'], row['col2']))
# conn.commit()
# cursor.close()
# conn.close()
# # 3. 從文件讀取與寫入
# # Excel 文件處理：

# # 使用 pandas 的 read_excel 和 to_excel 讀寫 Excel 文件。
# # 範例：

# df = pd.read_excel('your_file.xlsx')
# # 進行數據處理
# df.to_excel('output.xlsx', index=False)
# # Word 文件處理：

# # 使用 python-docx 讀取和寫入 Word 文件內容。
# # 範例：

# from docx import Document

# doc = Document('your_file.docx')
# for para in doc.paragraphs:
#     print(para.text)
# doc.save('output.docx')
# # CSV 文件處理：

# # 使用 pandas 的 read_csv 和 to_csv 讀寫 CSV 文件。
# # 範例：

# df = pd.read_csv('your_file.csv')
# df.to_csv('output.csv', index=False)
# # TXT 文件處理：

# # 使用 Python 的內建函數讀取和寫入 TXT 文件。
# # 範例：

# with open('your_file.txt', 'r') as file:
#     content = file.read()
#     print(content)
# with open('output.txt', 'w') as file:
#     file.write('Your content here')
# # 4. 數據集成與應用
# # 數據清洗與預處理：
# # 使用 pandas 進行數據清洗，如處理缺失值、數據格式轉換等。
# # 數據庫和文件之間的數據轉換：
# # 將來自不同文件格式的數據整理後批量導入到 MySQL，或者將數據從 MySQL 導出到文件。
# # 5. 實戰項目練習
# # 數據整合與視覺化：
# # 將從 MSSQL 和 MySQL 獲取的數據與文件數據進行整合，並使用 Matplotlib 或 Seaborn 進行視覺化展示。
# # 自動化數據處理：
# # 編寫腳本，定期從文件和數據庫中提取、處理並保存數據。
# # 6. 常見問題與解決方案
# # 連接問題：如何解決連接 MSSQL 或 MySQL 失敗的問題。
# # 性能優化：處理大文件和大數據時的性能優化技巧。
# # 這個大綱涵蓋了從資料庫串接到文件數據處理的整體流程，可以根據具體教學需求進行擴展和深入。


# # 要將 Planner 整合到 Python 工作流程中，使其可以直接從 Python 下語法來簡化作業流程，可以考慮以下步驟：

# # 1. 理解需求
# # 需求分析：確認你想要自動化的 Planner 作業內容，例如資料庫連接、資料查詢、數據處理、報表生成等。
# # 自動化目標：哪些具體的工作流程需要簡化？例如，是否需要定期從資料庫中提取數據並生成報表？是否需要自動化資料清洗和上傳？
# # 2. 設計 Python 腳本
# # 資料庫連接模組：
# # MSSQL：

# # 使用 Python 腳本來實現連接、查詢資料的功能，將 Planner 所需的查詢工作自動化。
# # 示例：

# import pyodbc

# def query_mssql(query):
#     conn = pyodbc.connect('DRIVER={SQL Server};SERVER=your_server;DATABASE=your_database;UID=your_user;PWD=your_password')
#     cursor = conn.cursor()
#     cursor.execute(query)
#     results = cursor.fetchall()
#     conn.close()
#     return results

# # 使用範例
# data = query_mssql("SELECT * FROM your_table")
# print(data)
# # MySQL：

# # 自動化資料庫操作，包括資料插入、更新和刪除。
# # 示例：

# import mysql.connector

# def insert_into_mysql(table, data):
#     conn = mysql.connector.connect(
#         host="10.96.196.36",
#         user="stan",
#         password="Ss31668984",
#         database="stan_test"
#     )
#     cursor = conn.cursor()
#     placeholders = ', '.join(['%s'] * len(data))
#     sql = f"INSERT INTO {table} VALUES ({placeholders})"
#     cursor.execute(sql, tuple(data.values()))
#     conn.commit()
#     cursor.close()
#     conn.close()

# # 使用範例
# insert_into_mysql("your_table", {"column1": value1, "column2": value2}) 
# # 3. 文件數據處理模組
# # 讀取與處理文件：
# # Excel、Word、CSV、TXT 文件的自動讀取、處理和轉換，將處理後的數據自動導入 MySQL。
# # 示例：

# import pandas as pd

# def process_csv_to_mysql(file_path, table_name):
#     df = pd.read_csv(file_path)
#     conn = mysql.connector.connect(
#         host="10.96.196.36",
#         user="stan",
#         password="Ss31668984",
#         database="stan_test"
#     )
#     cursor = conn.cursor()

#     for _, row in df.iterrows():
#         cursor.execute(f"INSERT INTO {table_name} (col1, col2) VALUES (%s, %s)", (row['col1'], row['col2']))
#     conn.commit()
#     cursor.close()
#     conn.close()

# # 使用範例
# process_csv_to_mysql('data.csv', 'your_table')
# # 4. 自動化 Planner 的功能
# # 作業排程：
# # 使用 schedule 或 cron 來定期執行 Python 腳本，實現自動化作業流程。例如，定期從 MSSQL 提取資料，處理後導入 MySQL，並生成報表。
# # 示例：

# import schedule
# import time

# def daily_task():
#     # 放置你要定期執行的代碼
#     data = query_mssql("SELECT * FROM your_table")
#     process_csv_to_mysql('data.csv', 'your_table')

# # 設置每天執行一次
# schedule.every().day.at("10:00").do(daily_task)

# while True:
#     schedule.run_pending()
#     time.sleep(60)  # 等待下一次執行
# # 5. 整合與測試
# # 整合腳本：將以上各個功能模組整合成一個完整的 Python 腳本或工具，以簡化操作流程。
# # 測試：進行多次測試，確保每個流程都能正確運行，並根據需要進行調整。
# # 6. 持續改進
# # 日誌記錄與錯誤處理：
# # 添加日誌記錄和錯誤處理機制，確保腳本在出現問題時能夠提示並記錄錯誤，方便調試和維護。
# # 優化與擴展：
# # 根據實際需求擴展腳本功能，如增加更多的資料處理、視覺化、報表生成功能。
# # 透過這種方式，可以將 Planner 的作業流程自動化和簡化，大幅提升工作效率並減少人為操作錯誤。



# # 以下是如何使用 Matplotlib、Seaborn 和 Plotly 進行數據可視化的學習和教學內容大綱：

# # 1. 數據可視化簡介
# # 數據可視化的重要性：
# # 探索性數據分析 (EDA) 的關鍵部分。
# # 協助發現數據模式、異常值和趨勢。
# # 工具簡介：
# # Matplotlib：強大的數據繪圖庫，適合基本和自定義可視化。
# # Seaborn：基於 Matplotlib，提供更高層次的界面，適合統計圖形和美觀的圖表。
# # Plotly：互動式可視化工具，適合網頁應用和動態展示。
# # 2. Matplotlib
# # 基礎圖表繪製：

# # 折線圖：

# import matplotlib.pyplot as plt

# x = [1, 2, 3, 4]
# y = [10, 20, 25, 30]

# plt.plot(x, y)
# plt.xlabel('X Axis')
# plt.ylabel('Y Axis')
# plt.title('Simple Line Plot')
# plt.show()
# # 柱狀圖：

# plt.bar(x, y)
# plt.xlabel('X Axis')
# plt.ylabel('Y Axis')
# plt.title('Simple Bar Plot')
# plt.show()
# # 散點圖：

# plt.scatter(x, y)
# plt.xlabel('X Axis')
# plt.ylabel('Y Axis')
# plt.title('Simple Scatter Plot')
# plt.show()
# # 圖表自定義：

# # 設置圖例、標籤、標題、顏色和線條樣式。
# # 調整子圖布局 (subplots) 和網格 (grid)。
# # 高級圖表：

# # 直方圖：

# import numpy as np

# data = np.random.randn(1000)
# plt.hist(data, bins=30, alpha=0.5)
# plt.title('Histogram')
# plt.show()
# # 箱線圖：

# data = np.random.rand(10, 4)
# plt.boxplot(data)
# plt.title('Boxplot')
# plt.show()
# # 3. Seaborn
# # 基本繪圖：

# # 分佈圖：

# import seaborn as sns
# import numpy as np

# data = np.random.randn(1000)
# sns.histplot(data, kde=True)
# plt.title('Histogram with KDE')
# plt.show()
# # 分類圖：

# tips = sns.load_dataset('tips')
# sns.boxplot(x='day', y='total_bill', data=tips)
# plt.title('Boxplot of Total Bill by Day')
# plt.show()
# # 關係圖：

# sns.scatterplot(x='total_bill', y='tip', data=tips)
# plt.title('Scatter Plot of Total Bill vs Tip')
# plt.show()
# # 進階繪圖：

# # 成對繪圖 (Pairplot)：

# # sns.pairplot(tips, hue='day')
# # plt.show()
# # 熱圖 (Heatmap)：

# flights = sns.load_dataset("flights")
# flights = flights.pivot("month", "year", "passengers")
# sns.heatmap(flights, annot=True, fmt="d", cmap="YlGnBu")
# plt.title('Heatmap of Flight Data')
# plt.show()
# # 樣式和主題：

# # 使用 Seaborn 的 set_style、set_palette 等方法設置圖表的美觀風格。
# # 4. Plotly
# # 安裝與介紹：

# # 安裝 Plotly 庫 (pip install plotly)，介紹 Plotly 的特點和應用場景。
# # 基本互動圖表：

# # 折線圖：

# import plotly.graph_objects as go

# fig = go.Figure(data=go.Scatter(x=x, y=y, mode='lines'))
# fig.show()
# # 柱狀圖：

# fig = go.Figure(data=[go.Bar(x=x, y=y)])
# fig.show()
# # 高級互動圖表：

# # 散點圖：

# fig = go.Figure(data=go.Scatter(x=x, y=y, mode='markers'))
# fig.show()
# # 3D 散點圖：

# z = [5, 15, 25, 35]
# fig = go.Figure(data=[go.Scatter3d(x=x, y=y, z=z, mode='markers')])
# fig.show()
# # 自定義與應用：

# # 如何自定義 Plotly 圖表，使用多圖層、注釋和滑桿等互動功能。
# # 整合 Plotly 圖表到網頁或 Jupyter Notebook 中進行展示。
# # 5. 數據可視化實戰
# # 實例分析：使用 Matplotlib、Seaborn 和 Plotly 進行數據集的可視化分析，包含從資料庫提取的數據和文件數據。
# # 互動可視化儀表板：使用 Plotly 和 Dash（Plotly 的 web 應用框架）來構建互動式儀表板。
# # 6. 總結與常見問題解答
# # 工具比較與選擇：根據需求選擇合適的可視化工具。
# # 常見問題與解決方案：解決圖表中可能遇到的顯示問題、性能問題和交互問題。
# # 這些內容可以幫助你在學習和教學中有效地使用 Matplotlib、Seaborn 和 Plotly 來進行數據可視化，滿足不同的數據展示需求。


# # 將 Planner 整合到資料處理流程中，可以讓你在 row data 時期就能看到資料的全貌，從而更迅速地檢查和分析資料。以下是如何實現這一點的具體方法：

# # 1. 數據全貌概覽
# # 資料預覽與檢查：

# # 使用 Python 來直接從資料庫或文件中提取資料，並利用可視化工具來初步檢查數據是否完整、有無異常。
# # 可以使用 Pandas 提供的功能，例如 head()、info()、describe() 等，來快速檢視數據的結構和統計摘要。

# import pandas as pd

# # 從資料庫或文件讀取數據
# df = pd.read_sql_query('SELECT * FROM your_table', connection)  # 從資料庫讀取
# # df = pd.read_csv('your_file.csv')  # 從CSV讀取

# # 檢視數據的前幾行
# print(df.head())

# # 檢視數據的基本信息
# print(df.info())

# # 檢視數據的統計摘要
# print(df.describe())
# # 2. 資料可視化
# # 使用 Matplotlib、Seaborn 或 Plotly：

# # 利用這些工具快速生成數據分佈圖、關係圖、趨勢圖等，幫助你更直觀地理解數據的結構和特徵。
# # 例如，可以使用 Seaborn 的 pairplot 來查看數據特徵之間的關係。

# import seaborn as sns
# import matplotlib.pyplot as plt

# # sns.pairplot(df)
# # plt.show()
# # 互動式檢查：

# # 使用 Plotly 來生成互動式圖表，這樣你可以通過縮放、滑動和點擊來更深入地探索數據。

# import plotly.express as px

# fig = px.scatter_matrix(df)
# fig.show()
# # 3. 快速數據檢查與清理
# # 資料清理：

# # 在檢視資料全貌的同時，可以使用 Python 自動化執行常見的資料清理任務，例如處理缺失值、重複數據、異常值等。
# # 可以通過 Pandas 的功能進行初步清理。

# # 處理缺失值
# df.dropna(inplace=True)  # 移除缺失值行
# # df.fillna(0, inplace=True)  # 用0替代缺失值

# # 處理重複數據
# df.drop_duplicates(inplace=True)

# # 處理異常值
# df = df[df['column_name'] < threshold]  # 移除異常值
# # 4. 資料分析與洞察
# # 數據分析：

# # 通過提前可視化和初步分析數據，可以更快地發現數據中的異常和趨勢，從而做出更明智的數據清理和處理決策。
# # 使用 Python 的統計和機器學習庫（如 SciPy、scikit-learn）進行更深入的數據分析，或使用聚合函數進行分組統計分析。

# # 分組統計
# group_data = df.groupby('category_column').mean()

# # 基本統計分析
# from scipy import stats

# t_statistic, p_value = stats.ttest_ind(df['group1'], df['group2'])
# print(f"T-statistic: {t_statistic}, P-value: {p_value}")
# # 5. 實時檢查與快速反饋
# # 自動化腳本：

# # 通過 Python 腳本，讓 Planner 在每次數據更新時自動運行上述的數據檢查和可視化流程，提供即時反饋。
# # 可以設置自動報告生成功能，在每次數據處理後生成 PDF 或 HTML 報告，總結關鍵發現。

# import pdfkit

# # 將分析結果保存為 HTML
# df.to_html('report.html')

# # 將 HTML 轉換為 PDF
# pdfkit.from_file('report.html', 'report.pdf')
# # 6. 整合與優化工作流程
# # 整合到現有工作流程：
# # 將這些數據檢查和分析功能整合到你現有的 Planner 工作流程中，使其在早期階段就能夠快速提供數據洞察。
# # 持續優化：
# # 根據使用需求和反饋，持續改進數據檢查和可視化流程，使之更加高效和準確。
# # 透過這些步驟，你可以在 row data 時期就全面掌握資料的情況，從而更快速地進行資料檢查與分析，提高整個資料處理流程的效率。

# import configparser
# import os
# import pandas as pd
# import chardet
# import json
# from pandas import json_normalize
 
# config = configparser.ConfigParser()
# config.read('config.ini')
 
# # Press the green button in the gutter to run the script.
# if __name__ == '__main__':
#     print(f'Program Start')
#     config_list_str = config['config_list']['config_list']
#     config_list = config_list_str.split(';')
#     print(f'config_list:{config_list}')
 
#     current_directory = os.path.dirname(os.path.realpath(__file__))
#     input_folder = os.path.join(current_directory, 'input')
#     output_folder = os.path.join(current_directory, 'output')
#     #
#     for conf in config_list:
#         file_name = config[conf]['file_name']
#         json_col = config[conf]['json_col']
#         print(f'------Start File Processing: {file_name}, KV in col: {json_col}-----')
 
#         file_path = os.path.join(input_folder, file_name)
 
#         # 偵測encoding
#         with open(file_path, 'rb') as file:
#             result = chardet.detect(file.read())
#         detected_encoding = result['encoding']
#         print(f"Detected encoding: {detected_encoding}")
 
#         # 讀檔
#         df = pd.read_csv(file_path, encoding=detected_encoding)
#         print('Read File Successfully')
 
#         df_normalized = json_normalize(df[json_col].apply(json.loads))
#         df = pd.concat([df, df_normalized], axis=1)
 
#         df.to_csv(f'{output_folder}\\{file_name}', index=False, encoding='utf-8-sig')
 
#         print('Save Output File Successfully')
 
#     print(f'Program End')

# -*- coding: utf-8 -*-
"""
Created on Mon May 22 11:06:41 2023

@author: Eric2_Ho
"""
import pandas as pd
import numpy as np
import json
import datetime
import pyodbc 
#從mssql中讀入新資料
# conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};'
#                       'SERVER=APZA001GOD;'
#                       'DATABASE=NPS;'
#                       'UID=NPSPO_User;'
#                       'PWD=Gaming@NPD')

# sql_query = ''' SELECT  * 
#                 FROM RawdataJoinKeywords2 
#                 where survey_version like '%2023%' '''
# df_new= pd.read_sql(sql_query, conn)
# conn.close()

#讀入新資料
df_new=pd.read_csv(r'D:\Eric工作\工作內容區\NPS\NPS每月VoC\20240328\NR\20240328-NR raw data.csv')
#讀入舊資料
df_old=pd.read_excel(r'D:\Eric工作\工作內容區\NPS\NPS每月VoC\20240229\NR\20240229-NR cleaned data.xlsx',sheet_name='Sheet1')
#%%df_new資料處理(為了讓new跟old格式切齊)
df_new=df_new.replace('',np.nan)
df_new = df_new.replace(['None', None], np.nan)
#%%找出舊資料中但未進行處理的response id
# df_old=df_old.drop(df_old[df_old['VOC_translate'].isnull()].index,axis=0)
#%%先把新增的資料切出來
dfnew=df_new[~df_new['Response_id'].isin(df_old['Response_id'])]
dfold=df_new[df_new['Response_id'].isin(df_old['Response_id'])]
#把新資料與舊資料打上不同的tag
dfnew['Response record']='this month'
dfold['Response record']='last month'
#聚合新舊資料
df=pd.concat([dfnew,dfold],axis=0)
#%%資料清理
# df_old=df_old.drop(['BU',
#        'EBS_Sales_Model_Name', 'ProductLevel1Name', 'ProductLevel2Name',
#        'productlevel3Name', 'SeriesName', 'Model_MKT_Name',
#        'month_since_Registration', 'Total_Registration', 'gender', 'age',
#        'Career', 'VOC_original', 
#        'keywords', 'function_team', 'category', 'event_', 'Model_Label'],axis=1)
#%%創造2個新欄位
#創造Answer Month
df['answer_time']=df['answer_time'].astype(str)
df['answer_month']=df['answer_time'].str.split('-').str.get(0)+'/'+df['answer_time'].str.split('-').str.get(1)
#創造Score Cluster
# 将 'Score' 列转换为数值类型，无法转换的值保持为 NaN
df['Score'] = pd.to_numeric(df['Score'], errors='coerce')
condi=[df['Score'].isnull(),(df['Score']>=0) & (df['Score']<=6),
       (df['Score']==7) | (df['Score']==8),(df['Score']==9) | (df['Score']==10)]
values=[np.nan,'Detractor','Passive','Promoter']
df['Score Cluster']=np.select(condi,values)
#%%欄位排序，最後整理
df=df.drop(['VOC_original','AsusSurvey_ID','keywords','function_team','category'],axis=1)
df=df[['survey_version','Response_id','Country', 'Score', 'answer_time',
       'VOC_translate','Model_Name', 'Segmentation',
       'Updated_keywords', 'Updated_function_team', 'Updated_category','Score Cluster','answer_month','Response record',
       'Territory','Project_Year_tg','source','Project_Year']]
#刪除model_name包含nr220的row
df=df[~df['Model_Name'].str.lower().str.contains('nr220',na=False)]
df=df.replace(np.nan,'Null')
#%%儲存
#製作日期檔名方便儲存
# 獲取當天的日期,並格式化成yyyymmdd
today = datetime.date.today().strftime('%Y%m%d')

#cleaned data
df=df.sort_values(['Response_id'],ascending=False)
#df.to_csv(fr'D:\Eric工作\工作內容區\NPS\NPS每月VoC\20230524-test\{today}-cleaned data.csv',index=False,encoding='utf-8-sig')
df.to_excel(fr'D:\Eric工作\工作內容區\NPS\NPS每月VoC\20240328\NR\{today}-NR cleaned data.xlsx',index=False,encoding='utf-8-sig')

#distinct cleaned data
df_d=df.drop(['Updated_keywords', 'Updated_function_team', 'Updated_category','Response record'],axis=1)
df_d=df_d.drop_duplicates()
df_d=df_d.sort_values(['Response_id'],ascending=False)
#df_d.to_csv(fr'D:\Eric工作\工作內容區\NPS\NPS每月VoC\20230524-test\{today}-cleaned data distinct.csv',index=False,encoding='utf-8-sig')
df_d.to_excel(fr'D:\Eric工作\工作內容區\NPS\NPS每月VoC\20240328\NR\{today}-NR cleaned data distinct.xlsx',index=False,encoding='utf-8-sig')
#%%
from openpyxl import load_workbook
from openpyxl.styles import Font, Color

# 加载现有的工作簿
wb = load_workbook(fr'D:\Eric工作\工作內容區\NPS\NPS每月VoC\20240328\NR\{today}-NR cleaned data.xlsx')
wb_d = load_workbook(fr'D:\Eric工作\工作內容區\NPS\NPS每月VoC\20240328\NR\{today}-NR cleaned data distinct.xlsx')
# 选择工作表
ws = wb.active
ws_d=wb_d.active
# 创建一个綠色字体
green_font = Font(color=Color('FF548235'))
# 創建一個空集合來收集改變顏色的行的id
changed_ids = set()
# 迭代工作表的所有行
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    # 检查a列的值是否为'本周'
    if row[13].value == 'this month':   # 注意：这里假设a列是第一列，如果不是，请相应地调整索引
        # 如果是，將此行的id添加到changed_ids集合中
        id = row[1].value  # 假設id是第二列，如果不是，請修改這裡的索引
        changed_ids.add(id)
        # 如果是，就把这一行的字体颜色改为我们定义的颜色
        for cell in row:
            cell.font = green_font
 # 迭代工作表的所有行           
for row in ws_d.iter_rows(min_row=2, max_row=ws_d.max_row, min_col=1, max_col=ws_d.max_column):

    # 檢查此行的id是否在changed_ids集合中
    id = row[1].value  # 假設id是第二列，如果不是，請修改這裡的索引
    if id in changed_ids:

        # 如果是，將此行的字體顏色更改為我們定義的顏色
        for cell in row:
            cell.font = green_font            
# 保存工作簿
wb.save(fr'D:\Eric工作\工作內容區\NPS\NPS每月VoC\20240328\NR\{today}-NR cleaned data.xlsx')
wb_d.save(fr'D:\Eric工作\工作內容區\NPS\NPS每月VoC\20240328\NR\{today}-NR cleaned data distinct.xlsx')

